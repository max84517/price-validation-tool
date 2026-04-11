"""
report/writer.py — write per-supplier Excel reports, one sheet per month.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from price_validation.config.paths import REPORT_DIR
from price_validation.validation.compare import MismatchRecord


# --------------------------------------------------------------------------- #
# Style constants
# --------------------------------------------------------------------------- #
_HEADER_FILL = PatternFill("solid", fgColor="2E4057")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_OK_FILL     = PatternFill("solid", fgColor="C6EFCE")
_WARN_FILL   = PatternFill("solid", fgColor="FFEB9C")
_ERR_FILL    = PatternFill("solid", fgColor="FFC7CE")

_FEATURE_HEADERS = [
    "HP/ODM Part#", "Color", "Product", "Size",
    "ODM & Site", "GTK Suppliers", "Platforms/Project",
]

_ALL_HEADERS = _FEATURE_HEADERS + [
    "Master Table Rebate", "Supplier Rebate", "Comment",
]


def _col_widths() -> list[int]:
    return [20, 12, 20, 8, 20, 20, 20, 14, 14, 60]


def _write_sheet(
    ws,
    month: str,
    records: list[MismatchRecord],
) -> None:
    # Header row
    for col_idx, header in enumerate(_ALL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    month_records = [r for r in records if r.month == month]

    if not month_records:
        # No mismatches — write success message
        cell = ws.cell(row=2, column=1, value="No mismatches found for this month.")
        cell.fill = _OK_FILL
        cell.font = Font(bold=True, color="375623", size=11)
        ws.column_dimensions["A"].width = 50
        return

    for row_idx, rec in enumerate(month_records, start=2):
        values = [
            rec.hp_odm_part, rec.color, rec.product, rec.size,
            rec.odm_site, rec.gtk_suppliers, rec.platforms,
            rec.pt_rebate, rec.shp_rebate, rec.comment,
        ]
        if not rec.exists_in_pt or not rec.exists_in_shp:
            fill = _WARN_FILL
        else:
            fill = _ERR_FILL

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col_idx, width in enumerate(_col_widths(), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


def write_report(
    supplier_name: str,
    fy: str,
    months: list[str],
    records: list[MismatchRecord],
) -> Path:
    """
    Create one Excel workbook for *supplier_name* with one sheet per month.
    Files are stored under data/report/<UTC timestamp>/.
    Returns the path of the written file.
    """
    now_local = datetime.now()
    folder_name = now_local.strftime("%Y-%m-%d %H:%M")
    out_dir = REPORT_DIR / folder_name
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in supplier_name)
    filename = out_dir / f"Report_{safe_name}_FY{fy}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for month in months:
        ws = wb.create_sheet(title=month)
        _write_sheet(ws, month, records)

    wb.save(filename)
    return filename

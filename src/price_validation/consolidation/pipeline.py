"""
consolidation/pipeline.py
=========================
Full pipeline: ingest raw → consolidate by segment → consolidate all
→ rebate only → write to Pricing_Template_InputDevices.xlsx InputDevice sheet.

Adapted from master-price-consolidate (github.com/max84517/master-price-consolidate).

Source folder structure expected from the user:
  nb_kb/
    Master price table_bNB/
      Master price table_bNB_<SUPPLIER>/   ← newest .xlsx is picked
      ...
    Master price table_cNB/
      Master price table_cNB_<SUPPLIER>/
      ...
  dt_kb/
    Master price table_DT/
      Master price table_DT_<SUPPLIER>/
      ...
  peripheral/
    Master price table_Peripheral/
      Master price table_Peripheral_<SUPPLIER>/
      ...
"""
from __future__ import annotations

import re
import tempfile
from datetime import date
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl import Workbook

from price_validation.config.paths import PRICING_TEMPLATE_DIR, MASTER_PRICE_SOURCE_DIR

# ── Constants ─────────────────────────────────────────────────────────────────

_FY_PATTERN = re.compile(r"^FY\d{2}$")
_REMOVE_PATTERNS = ("HP Cost", "ODM Cost")

# Matches new combined NB sheet names: "FY25 cNB", "FY25 bNB", "fy26CNB", "FY25 c NB"
# Captures (fy_year_digits, 'c'|'b'). Must be the entire string (no extra text).
_NB_SHEET_RE = re.compile(r"^FY(\d{2})\s*([cb])\s*NB$", re.IGNORECASE)

# NB combined folder constants
_NB_MASTER_SUB = "Master price table_NB"
_NB_PREFIX     = "Master price table_NB_"

# Non-NB segments only (bNB and cNB are now handled by _ingest_nb)
_SEGMENT_DEFS: list[tuple[str, str, str, str]] = [
    ("DT",          "dt_kb",        "Master price table_DT",         "Master price table_DT_"),
    ("Peripheral",  "peripheral",   "Master price table_Peripheral", "Master price table_Peripheral_"),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_fy_sheet(name: str) -> bool:
    return bool(_FY_PATTERN.match(name.strip()))


def _newest_xlsx(folder: Path) -> Optional[Path]:
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() == ".xlsx"]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def _valid_col_indices(src_sheet) -> list[int]:
    """Header is on row 2; return indices of columns that have a non-None header."""
    rows = list(src_sheet.iter_rows(values_only=True, min_row=2, max_row=2))
    if not rows:
        return []
    return [i for i, h in enumerate(rows[0]) if h is not None]


def _copy_rows_segment(src_sheet, dest_sheet, first_supplier: bool, col_indices: list[int]) -> None:
    """
    Source layout: row 1 = title (skip), row 2 = headers (skip if not first), row 3+ = data.
    Normalise header cells: replace \\n with space.
    Only include columns in col_indices.
    """
    for row_idx, row in enumerate(src_sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        if row_idx == 2 and not first_supplier:
            continue
        extracted = [row[i] if i < len(row) else None for i in col_indices]
        if row_idx == 2:
            extracted = [v.replace("\n", " ") if isinstance(v, str) else v for v in extracted]
        if all(c is None for c in extracted):
            continue
        dest_sheet.append(extracted)


def _should_remove_col(header: object) -> bool:
    if not isinstance(header, str):
        return False
    h = header.replace("\n", " ")
    return any(pat.lower() in h.lower() for pat in _REMOVE_PATTERNS)


# ── Stage 1: Ingest ───────────────────────────────────────────────────────────

def _fix_gtk_suppliers(
    xlsx_path: Path,
    supplier_name: str,
    log: Callable[[str, str], None],
    seg_label: str,
) -> None:
    """
    Open *xlsx_path*, find the 'GTK Suppliers' column (header on row 2),
    and overwrite every data-row cell with *supplier_name*.
    This prevents mis-labelled or empty GTK Suppliers values from breaking
    the per-supplier split during validation.
    """
    # data_only=True: formula cells are loaded as their cached values (literals).
    # This prevents openpyxl from stripping cached values on save, which would
    # cause consolidation (also data_only=True) to read those cells as None.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        if not _is_fy_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        # Header is on row 2; row 1 is a title row
        header_row = [c.value for c in ws[2]]
        gtk_col_idx = None
        for i, h in enumerate(header_row):
            if isinstance(h, str) and h.strip().lower() == "gtk suppliers":
                gtk_col_idx = i + 1  # openpyxl is 1-based
                break
        if gtk_col_idx is None:
            log(f"[{seg_label}] 'GTK Suppliers' column not found in {xlsx_path.name} sheet {sheet_name}", "WARN")
            continue
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            row[gtk_col_idx - 1].value = supplier_name
    wb.save(xlsx_path)
    wb.close()


def _ingest_to_dir(
    source_paths: dict[str, str],
    raw_dir: Path,
    log: Callable[[str, str], None],
) -> dict[str, list[Path]]:
    """
    Copy newest .xlsx from each supplier sub-folder into raw_dir/<segment>/.
    Returns {segment_label: [list of copied paths]}.
    """
    result: dict[str, list[Path]] = {}

    for seg_label, src_key, master_sub, prefix in _SEGMENT_DEFS:
        master_dir = Path(source_paths[src_key]) / master_sub
        if not master_dir.exists():
            log(f"[{seg_label}] Folder not found: {master_dir}", "WARN")
            result[seg_label] = []
            continue

        seg_raw_dir = raw_dir / seg_label
        seg_raw_dir.mkdir(parents=True, exist_ok=True)
        copied: list[Path] = []

        matched = sorted(d for d in master_dir.iterdir() if d.is_dir() and d.name.startswith(prefix))
        if not matched:
            log(f"[{seg_label}] No sub-folders matching '{prefix}*' in {master_dir}", "WARN")
            result[seg_label] = []
            continue

        for supplier_dir in matched:
            latest = _newest_xlsx(supplier_dir)
            if latest is None:
                log(f"[{seg_label}] No .xlsx in {supplier_dir.name}", "WARN")
                continue
            short = supplier_dir.name.replace("Master price table_", "", 1)
            dest = seg_raw_dir / f"{short}.xlsx"
            import shutil
            shutil.copy2(latest, dest)
            # ── Overwrite GTK Suppliers column with supplier name from filename ──
            # short is like "DT_CHICONY" or "bNB_LITEON"; supplier = part after first "_"
            supplier_name_from_file = short.split("_", 1)[1] if "_" in short else short
            _fix_gtk_suppliers(dest, supplier_name_from_file, log, seg_label)
            # ── Copy to master price source folder (keep original filename) ──
            master_src_seg_dir = MASTER_PRICE_SOURCE_DIR / seg_label
            shutil.copy2(dest, master_src_seg_dir / latest.name)
            copied.append(dest)
            log(f"[{seg_label}] Ingested: {supplier_dir.name}/{latest.name}", "INFO")

        result[seg_label] = copied

    return result


# ── Stage 1b: Ingest NB combined folder ─────────────────────────────────────

def _ingest_nb(
    source_paths: dict[str, str],
    raw_dir: Path,
    log: Callable[[str, str], None],
) -> dict[str, list[Path]]:
    """
    Handle the new combined NB source structure:
      <nb_kb>/Master price table_NB/Master price table_NB_<Supplier>/*.xlsx

    Each supplier Excel contains sheets named 'FY25 cNB', 'FY25 bNB', etc.
    This function:
      1. Scans all supplier Excels to find FY years that have BOTH cNB and bNB
         sheets present (across all supplier files combined).
      2. For each supplier, writes two output Excels:
         - raw_dir/bNB/<supplier>.xlsx  — only bNB sheets, renamed to FY{YY}
         - raw_dir/cNB/<supplier>.xlsx  — only cNB sheets, renamed to FY{YY}
      3. Applies _fix_gtk_suppliers and copies to master price source.
    """
    import shutil

    nb_kb_path = Path(source_paths.get("nb_kb", ""))
    master_dir = nb_kb_path / _NB_MASTER_SUB
    if not master_dir.exists():
        log(f"[NB] Folder not found: {master_dir}", "WARN")
        return {"bNB": [], "cNB": []}

    supplier_dirs = sorted(
        d for d in master_dir.iterdir()
        if d.is_dir() and d.name.startswith(_NB_PREFIX)
    )
    if not supplier_dirs:
        log(f"[NB] No sub-folders matching '{_NB_PREFIX}*' in {master_dir}", "WARN")
        return {"bNB": [], "cNB": []}

    # ── Pass 1: collect valid FY years (must have both cNB and bNB across all suppliers) ──
    cnb_fys: set[str] = set()
    bnb_fys: set[str] = set()
    for supplier_dir in supplier_dirs:
        latest = _newest_xlsx(supplier_dir)
        if latest is None:
            continue
        wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            m = _NB_SHEET_RE.match(sheet_name.strip())
            if m:
                fy = f"FY{m.group(1)}"
                if m.group(2).lower() == "c":
                    cnb_fys.add(fy)
                else:
                    bnb_fys.add(fy)
        wb.close()

    valid_fys = cnb_fys & bnb_fys
    if not valid_fys:
        log("[NB] No FY year has both cNB and bNB sheets — skipping NB ingest", "WARN")
        return {"bNB": [], "cNB": []}
    log(f"[NB] Valid FY years (cNB+bNB both present): {sorted(valid_fys)}", "INFO")

    # ── Pass 2: split each supplier Excel into bNB and cNB files ──
    bnb_raw = raw_dir / "bNB"
    cnb_raw = raw_dir / "cNB"
    bnb_raw.mkdir(parents=True, exist_ok=True)
    cnb_raw.mkdir(parents=True, exist_ok=True)

    result: dict[str, list[Path]] = {"bNB": [], "cNB": []}

    for supplier_dir in supplier_dirs:
        latest = _newest_xlsx(supplier_dir)
        if latest is None:
            log(f"[NB] No .xlsx in {supplier_dir.name}", "WARN")
            continue

        short = supplier_dir.name.replace("Master price table_", "", 1)  # e.g. "NB_CHICONY"
        supplier_name = short.split("_", 1)[1] if "_" in short else short

        src_wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
        # raw workbooks: normalised sheet names (FY25) — used by pipeline
        bnb_out = Workbook(); bnb_out.remove(bnb_out.active)
        cnb_out = Workbook(); cnb_out.remove(cnb_out.active)
        # master workbooks: original sheet names (FY25 bNB) — saved to master price source
        bnb_master = Workbook(); bnb_master.remove(bnb_master.active)
        cnb_master = Workbook(); cnb_master.remove(cnb_master.active)
        bnb_wrote = False
        cnb_wrote = False

        for sheet_name in src_wb.sheetnames:
            m = _NB_SHEET_RE.match(sheet_name.strip())
            if not m:
                continue
            fy = f"FY{m.group(1)}"
            if fy not in valid_fys:
                continue
            nb_type = m.group(2).lower()  # 'b' or 'c'
            target_raw    = bnb_out    if nb_type == "b" else cnb_out
            target_master = bnb_master if nb_type == "b" else cnb_master
            src_sheet = src_wb[sheet_name]
            rows = [list(r) for r in src_sheet.iter_rows(values_only=True)]
            # raw: normalised name for downstream processing
            raw_sheet = target_raw.create_sheet(title=fy)
            for row in rows:
                raw_sheet.append(row)
            # master: keep original sheet name
            master_sheet = target_master.create_sheet(title=sheet_name.strip())
            for row in rows:
                master_sheet.append(row)
            if nb_type == "b":
                bnb_wrote = True
            else:
                cnb_wrote = True

        src_wb.close()

        for seg_label, out_wb, master_wb, wrote, seg_raw in [
            ("bNB", bnb_out, bnb_master, bnb_wrote, bnb_raw),
            ("cNB", cnb_out, cnb_master, cnb_wrote, cnb_raw),
        ]:
            if not wrote:
                log(f"[{seg_label}] No valid sheets for {supplier_dir.name} — skipped", "WARN")
                continue
            dest = seg_raw / f"{short}.xlsx"
            out_wb.save(dest)
            _fix_gtk_suppliers(dest, supplier_name, log, seg_label)
            master_src_seg_dir = MASTER_PRICE_SOURCE_DIR / seg_label
            master_dest = master_src_seg_dir / latest.name
            master_wb.save(master_dest)
            result[seg_label].append(dest)
            log(f"[{seg_label}] Ingested: {supplier_dir.name}/{latest.name}", "INFO")

    return result


# ── Stage 2: Consolidate by segment ──────────────────────────────────────────

def _consolidate_segment(
    seg_label: str,
    raw_folder: Path,
    out_dir: Path,
    log: Callable[[str, str], None],
) -> Optional[Path]:
    xlsx_files = sorted(raw_folder.glob("*.xlsx"))
    if not xlsx_files:
        log(f"[{seg_label}] No raw files to consolidate", "WARN")
        return None

    today = date.today().strftime("%Y%m%d")
    out_path = out_dir / f"Consolidated_{seg_label}_{today}.xlsx"

    fy_order: list[str] = []
    fy_sources: dict[str, list[tuple[Path, str]]] = {}

    for xlsx_path in xlsx_files:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if not _is_fy_sheet(sheet_name):
                continue
            fy = sheet_name.strip()
            if fy not in fy_sources:
                fy_order.append(fy)
                fy_sources[fy] = []
            fy_sources[fy].append((xlsx_path, sheet_name))
        wb.close()

    if not fy_order:
        log(f"[{seg_label}] No FY sheets found", "WARN")
        return None

    fy_order.sort()
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for fy in fy_order:
        out_sheet = out_wb.create_sheet(title=fy)
        first = True
        col_indices: list[int] = []
        for xlsx_path, sheet_name in fy_sources[fy]:
            src_wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            src_sheet = src_wb[sheet_name]
            if first:
                col_indices = _valid_col_indices(src_sheet)
            _copy_rows_segment(src_sheet, out_sheet, first_supplier=first, col_indices=col_indices)
            src_wb.close()
            first = False
        log(f"[{seg_label}] {fy}: {len(fy_sources[fy])} supplier(s), {out_sheet.max_row} rows", "INFO")

    out_wb.save(out_path)
    return out_path


# ── Stage 3: Consolidate all segments ────────────────────────────────────────

def _consolidate_all(
    segment_files: list[Path],
    out_dir: Path,
    log: Callable[[str, str], None],
) -> Path:
    """Stack segment files by FY sheet. Header is row 1 in segment output."""
    fy_order: list[str] = []
    fy_sources: dict[str, list[tuple[Path, str]]] = {}

    for src_path in segment_files:
        wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name not in fy_sources:
                fy_order.append(sheet_name)
                fy_sources[sheet_name] = []
            fy_sources[sheet_name].append((src_path, sheet_name))
        wb.close()

    fy_order.sort()
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for fy in fy_order:
        out_sheet = out_wb.create_sheet(title=fy)
        first = True
        for src_path, sheet_name in fy_sources[fy]:
            src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
            src_sheet = src_wb[sheet_name]
            for row_idx, row in enumerate(src_sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1 and not first:
                    continue
                if all(c is None for c in row):
                    continue
                out_sheet.append(list(row))
            src_wb.close()
            first = False
        log(f"[All] {fy}: {len(fy_sources[fy])} segment(s), {out_sheet.max_row} rows", "INFO")

    today = date.today().strftime("%Y%m%d")
    out_path = out_dir / f"consolidate_all_{today}.xlsx"
    out_wb.save(out_path)
    return out_path


# ── Stage 4: Rebate only ─────────────────────────────────────────────────────

def _rebate_only(all_path: Path, out_dir: Path) -> Path:
    src_wb = openpyxl.load_workbook(all_path, read_only=True, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for sheet_name in src_wb.sheetnames:
        src_sheet = src_wb[sheet_name]
        out_sheet = out_wb.create_sheet(title=sheet_name)
        header_row = next(src_sheet.iter_rows(values_only=True), ())
        keep = [i for i, h in enumerate(header_row) if not _should_remove_col(h)]
        for row_idx, row in enumerate(src_sheet.iter_rows(values_only=True), start=1):
            extracted = [row[i] if i < len(row) else None for i in keep]
            if row_idx > 1 and all(c is None for c in extracted):
                continue
            out_sheet.append(extracted)

    src_wb.close()
    out_path = out_dir / "rebate_only.xlsx"
    out_wb.save(out_path)
    return out_path


# ── Stage 5: Write to Pricing Template ───────────────────────────────────────

def _write_pricing_template(rebate_path: Path, fy_sheet: str) -> Path:
    """
    Write a fresh Pricing_Template_InputDevices.xlsx with a single sheet
    named after the selected FY (e.g. 'FY25'). Overwrites any existing file.
    """
    PRICING_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    template_file = PRICING_TEMPLATE_DIR / "Pricing_Template_InputDevices.xlsx"

    src_wb = openpyxl.load_workbook(rebate_path, read_only=True, data_only=True)
    if fy_sheet not in src_wb.sheetnames:
        src_wb.close()
        raise ValueError(f"Sheet '{fy_sheet}' not found in rebate workbook")
    src_sheet = src_wb[fy_sheet]

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    dest_sheet = out_wb.create_sheet(title=fy_sheet)

    # Locate the Platforms/Project column index from the header row
    header: tuple = ()
    platforms_col_idx: int = -1
    rows_iter = src_sheet.iter_rows(values_only=True)
    for row in rows_iter:
        header = row
        # Normalise header cell: collapse whitespace / newlines
        platforms_col_idx = next(
            (i for i, h in enumerate(header)
             if isinstance(h, str) and re.sub(r'\s+', ' ', h).strip().lower() == "platforms/project"),
            -1,
        )
        dest_sheet.append(list(header))
        break  # only process header row here

    for row in rows_iter:
        if all(c is None for c in row):
            continue
        # Skip rows where Platforms/Project is blank
        if platforms_col_idx >= 0:
            val = row[platforms_col_idx] if platforms_col_idx < len(row) else None
            if val is None or str(val).strip() == "":
                continue
        dest_sheet.append(list(row))

    src_wb.close()
    out_wb.save(template_file)
    return template_file


# ── Public API ────────────────────────────────────────────────────────────────

def check_latest_files(source_paths: dict[str, str]) -> list[dict]:
    """
    For each segment × supplier folder, find the newest .xlsx file and return
    its metadata without copying or modifying anything.

    Returns a list of dicts, one per supplier folder found:
        {
            "segment":   str,          # e.g. "NB", "DT", "Peripheral"
            "supplier":  str,          # folder name suffix
            "filename":  str | None,   # latest .xlsx filename, or None if empty
            "modified":  datetime | None,
        }
    """
    from datetime import datetime
    results: list[dict] = []

    # ── NB combined folder ──
    nb_kb_path = Path(source_paths.get("nb_kb", ""))
    nb_master_dir = nb_kb_path / _NB_MASTER_SUB
    if nb_master_dir.exists():
        matched_nb = sorted(
            d for d in nb_master_dir.iterdir()
            if d.is_dir() and d.name.startswith(_NB_PREFIX)
        )
        for supplier_dir in matched_nb:
            short = supplier_dir.name.replace("Master price table_", "", 1)
            latest = _newest_xlsx(supplier_dir)
            results.append({
                "segment":  "NB",
                "supplier": short,
                "filename": latest.name if latest else None,
                "modified": datetime.fromtimestamp(latest.stat().st_mtime) if latest else None,
            })

    # ── Non-NB segments (DT, Peripheral) ──
    for seg_label, src_key, master_sub, prefix in _SEGMENT_DEFS:
        master_dir = Path(source_paths[src_key]) / master_sub
        if not master_dir.exists():
            continue
        matched = sorted(d for d in master_dir.iterdir() if d.is_dir() and d.name.startswith(prefix))
        for supplier_dir in matched:
            short = supplier_dir.name.replace("Master price table_", "", 1)
            latest = _newest_xlsx(supplier_dir)
            results.append({
                "segment":  seg_label,
                "supplier": short,
                "filename": latest.name if latest else None,
                "modified": datetime.fromtimestamp(latest.stat().st_mtime) if latest else None,
            })

    return results


def check_missing_fy_sheets(source_paths: dict[str, str], fy_sheet: str) -> list[dict]:
    """
    For the given FY sheet name (e.g. 'FY25'), inspect each supplier folder
    and return a list of dicts for suppliers that are missing required sheets.

    Each dict: {"segment": str, "supplier": str, "missing": list[str]}

    For NB suppliers: looks for 'FY25 bNB' and 'FY25 cNB' sheets.
    For DT / Peripheral: looks for 'FY25' sheet directly.
    """
    results: list[dict] = []

    if not _FY_PATTERN.match(fy_sheet.strip()):
        return []
    year_digits = fy_sheet.strip()[2:]  # "FY25" -> "25"

    # ── NB ──
    nb_kb_path = Path(source_paths.get("nb_kb", ""))
    nb_master_dir = nb_kb_path / _NB_MASTER_SUB
    if nb_master_dir.exists():
        for supplier_dir in sorted(
            d for d in nb_master_dir.iterdir()
            if d.is_dir() and d.name.startswith(_NB_PREFIX)
        ):
            short = supplier_dir.name.replace("Master price table_", "", 1)
            latest = _newest_xlsx(supplier_dir)
            if latest is None:
                results.append({"segment": "NB", "supplier": short,
                                 "missing": [f"FY{year_digits} bNB", f"FY{year_digits} cNB"]})
                continue
            wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
            stripped = {s.strip() for s in wb.sheetnames}
            wb.close()
            found_b = any(
                (m := _NB_SHEET_RE.match(s)) and m.group(1) == year_digits and m.group(2).lower() == "b"
                for s in stripped
            )
            found_c = any(
                (m := _NB_SHEET_RE.match(s)) and m.group(1) == year_digits and m.group(2).lower() == "c"
                for s in stripped
            )
            missing = []
            if not found_b:
                missing.append(f"FY{year_digits} bNB")
            if not found_c:
                missing.append(f"FY{year_digits} cNB")
            if missing:
                results.append({"segment": "NB", "supplier": short, "missing": missing})

    # ── DT / Peripheral ──
    for seg_label, src_key, master_sub, prefix in _SEGMENT_DEFS:
        master_dir = Path(source_paths[src_key]) / master_sub
        if not master_dir.exists():
            continue
        for supplier_dir in sorted(
            d for d in master_dir.iterdir()
            if d.is_dir() and d.name.startswith(prefix)
        ):
            short = supplier_dir.name.replace("Master price table_", "", 1)
            latest = _newest_xlsx(supplier_dir)
            if latest is None:
                results.append({"segment": seg_label, "supplier": short, "missing": [fy_sheet]})
                continue
            wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
            sheet_names = {s.strip() for s in wb.sheetnames}
            wb.close()
            if fy_sheet not in sheet_names:
                results.append({"segment": seg_label, "supplier": short, "missing": [fy_sheet]})

    return results


def get_available_fy_sheets(source_paths: dict[str, str], log: Callable[[str, str], None]) -> list[str]:
    """
    Run ingest + segment consolidation + consolidate_all in a temp dir
    and return the list of available FY sheet names.
    Stores the temp dir path internally so run_full_pipeline can reuse it.
    """
    # We do a dry run to discover sheets; the results are cached in a module-level temp dir
    global _cached_rebate_path, _cached_tmpdir
    _cleanup_cache()
    _cached_tmpdir = tempfile.mkdtemp(prefix="pvtool_build_")
    tmp = Path(_cached_tmpdir)

    _cached_rebate_path = _run_stages(source_paths, tmp, log)
    if _cached_rebate_path is None:
        return []

    wb = openpyxl.load_workbook(_cached_rebate_path, read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    return sheets


def run_full_pipeline(
    source_paths: dict[str, str],
    fy_sheet: str,
    log: Callable[[str, str], None],
) -> Path:
    """
    Run the full pipeline and write the chosen FY sheet to the pricing template.
    If get_available_fy_sheets was called first and the source paths are the same,
    reuses the cached rebate workbook. Otherwise re-runs from scratch.
    """
    global _cached_rebate_path, _cached_tmpdir
    if _cached_rebate_path is None or not _cached_rebate_path.exists():
        _cleanup_cache()
        _cached_tmpdir = tempfile.mkdtemp(prefix="pvtool_build_")
        tmp = Path(_cached_tmpdir)
        _cached_rebate_path = _run_stages(source_paths, tmp, log)

    if _cached_rebate_path is None:
        raise RuntimeError("Pipeline failed — check log for details.")

    out = _write_pricing_template(_cached_rebate_path, fy_sheet)
    log(f"Pricing Template written: {out}", "SUCCESS")
    _cleanup_cache()
    return out


def _run_stages(
    source_paths: dict[str, str],
    tmp: Path,
    log: Callable[[str, str], None],
) -> Optional[Path]:
    """Run ingest → segment → all → rebate_only; return rebate path or None on failure."""
    import shutil as _shutil

    # ── Clear master price source folders before each run ──
    _MASTER_SEGS = ("bNB", "cNB", "DT", "Peripheral")
    for seg in _MASTER_SEGS:
        seg_dir_ms = MASTER_PRICE_SOURCE_DIR / seg
        seg_dir_ms.mkdir(parents=True, exist_ok=True)
        for f in seg_dir_ms.glob("*.xlsx"):
            f.unlink(missing_ok=True)

    raw_dir = tmp / "raw"
    seg_dir = tmp / "segment"
    all_dir = tmp / "all"
    rebate_dir = tmp / "rebate"

    for d in (raw_dir, seg_dir, all_dir, rebate_dir):
        d.mkdir(parents=True, exist_ok=True)

    log("Ingesting raw files…", "INFO")
    _ingest_nb(source_paths, raw_dir, log)
    _ingest_to_dir(source_paths, raw_dir, log)

    # Consolidate all 4 segments (bNB and cNB come from _ingest_nb, DT/Peripheral from _ingest_to_dir)
    _ALL_SEG_LABELS = ("bNB", "cNB") + tuple(s[0] for s in _SEGMENT_DEFS)
    segment_files: list[Path] = []
    for seg_label in _ALL_SEG_LABELS:
        seg_raw = raw_dir / seg_label
        if not seg_raw.exists() or not any(seg_raw.glob("*.xlsx")):
            log(f"[{seg_label}] Skipped (no raw files)", "WARN")
            continue
        log(f"Consolidating segment {seg_label}…", "INFO")
        seg_out = _consolidate_segment(seg_label, seg_raw, seg_dir, log)
        if seg_out:
            segment_files.append(seg_out)

    if not segment_files:
        log("No segment files produced — aborting.", "ERROR")
        return None

    log("Consolidating all segments…", "INFO")
    all_path = _consolidate_all(segment_files, all_dir, log)
    log("Building rebate-only workbook…", "INFO")
    rebate_path = _rebate_only(all_path, rebate_dir)
    return rebate_path


# ── Cache state ───────────────────────────────────────────────────────────────

_cached_rebate_path: Optional[Path] = None
_cached_tmpdir: Optional[str] = None


def _cleanup_cache() -> None:
    global _cached_rebate_path, _cached_tmpdir
    if _cached_tmpdir:
        import shutil
        try:
            shutil.rmtree(_cached_tmpdir, ignore_errors=True)
        except Exception:
            pass
    _cached_rebate_path = None
    _cached_tmpdir = None
